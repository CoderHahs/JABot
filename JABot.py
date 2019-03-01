


'''
Setting up a secure, encrypted SMTP connection.

creates a secure connection with Gmail’s SMTP server, using the SMTP_SSL() 
of smtplib to initiate a TLS-encrypted connection. The default context of ssl validates the host name and its certificates and optimizes the security of the connection. 
'''

hrithik@facebook.com
shah@facebook.com
hrithikshah@facebook.com
hrithik.shah@facebook.com
hshah@facebook.com
h.shah@facebook.com
hrithiks@facebook.com
hrithik.s@facebook.com
hs@facebook.com
h.s@facebook.com
shahhrithik@facebook.com
shah.hrithik@facebook.com
shahh@facebook.com
shah.h@facebook.com
shrithik@facebook.com
s.hrithik@facebook.com
sh@facebook.com
s.h@facebook.com
hrithik-shah@facebook.com
h-shah@facebook.com
hrithik-s@facebook.com
h-s@facebook.com
shah-hrithik@facebook.com
shah-h@facebook.com
s-hrithik@facebook.com
s-h@facebook.com
hrithik_shah@facebook.com
h_shah@facebook.com
hrithik_s@facebook.com
h_s@facebook.com
shah_hrithik@facebook.com
shah_h@facebook.com
s_hrithik@facebook.com
s_h@facebook.com

import smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import xlrd # excel manipulations
from datetime import datetime 
import openpyxl # excel manipulations
import getpass #secure password input
#from email_sender import send_mail
#from VerifyEmailAddress import verifier
'''import re
import smtplib
import dns.resolver'''

port = 465  # For SSL
#print('hi')
sender_email = "hrithikshah00@gmail.com"
print (sender_email)
password = getpass.getpass("Type your password and press enter: ")
print ("Starting emailing....")

  
  


excelwb = "googleunirecruiters.xlsx"

# open book
book = xlrd.open_workbook(excelwb)
sh = book.sheet_by_index(0)
wb = openpyxl.load_workbook(filename = excelwb)
ws = wb.worksheets[0]

row = int(sh.cell_value(rowx=1, colx=13))
print(row)



while (row < sh.nrows):
  # who to send a message to
  firstName = str(sh.cell_value(rowx=row, colx=2))
  lastName = str(sh.cell_value(rowx=row, colx=3))
  at_index = (str(sh.cell_value(rowx=row, colx=4)).find(" at "))
  #company = str(sh.cell_value(rowx=row, colx=4))[at_index+4:]
  company = "google"

  combination0 = firstName[0]+lastName+"@"+company+".com"
  combination1 = firstName+"."+lastName+"@"+company+".com"
  combination2 = firstName+lastName+"@"+company+".com"
  combination3 = lastName+"@"+company+".com"
  combination4 = firstName+"_"+lastName+"@"+company+".com"
  combination5 = firstName[0]+"_"+lastName+"@"+company+".com"
  combination6 = firstName+lastName[0]+"@"+company+".com"
  combination7 = firstName+"@"+company+".com"
  combination8 = lastName+firstName[0]+"@"+company+".com"
  combination9 = lastName+"."+firstName+"@"+company+".com"
  combination10 = lastName[0]+firstName+"@"+company+".com"
  combination11 = lastName+firstName+"@"+company+".com"
  email_combinations = {"0":combination0, "1":combination1, "2": combination2, "3": combination3, "4": combination4, "5": combination5, "6": combination6, "7": combination7, "8": combination8, "9": combination9, "10": combination10,"11": combination11}

  if (at_index == -1):
    row += 1
    continue

  # try various emails
  for i in range(len(email_combinations)):
    receiver_email = (email_combinations[str(i)].lower())
    #print (receiver_email)
    # message stuff
    message = MIMEMultipart()
    message["Subject"] = "Application for Co-op Position Summer 2019"
    message["From"] = "Hrithik Shah"
    message["To"] = receiver_email

    html = """\
        <html>
          <body>
            <div class="">
           <div class="aHl"></div>
           <div id=":16i" tabindex="-1"></div>
           <div id=":16v" class="ii gt">
              <div id=":16w" class="a3s aXjCH ">
                 <div dir="ltr">
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif;font-size:11pt">Dear """+firstName+" "+lastName+""",</span><br></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">&nbsp;</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">I hope you have had a wonderful start to the new year!</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">&nbsp;</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">My name is Hrithik Shah and I am an undergrad student pursuing&nbsp;<strong>Software Engineering (CO-OP)</strong>&nbsp;at the University of Ottawa. As part of my undergraduate program I am looking to complete an internship for&nbsp;<strong>Summer 2019</strong>&nbsp;and am pleased&nbsp;</span><span style="font-family:&quot;Calibri Light&quot;,sans-serif;color:black">to present to you my resume (attached with this email) as consideration for the following role(s):</span></p>
                    <ul>
                       <li style="margin-left:15px"><span style="color:rgb(0,0,0);font-family:&quot;Calibri Light&quot;,sans-serif"><span style="font-size:14.6667px"><strong>Software Developer Intern</strong></span></span></li>
                       <li style="margin-left:15px"><font color="#000000" face="Calibri Light, sans-serif"><span style="font-size:14.6667px"><b>Data Analyst Intern</b></span></font></li>
                    </ul>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">As someone who enjoys technology and working on software development projects, I believe I would learn a lot through an internship and would prove to be an immediate contributor to your firm.</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">&nbsp;</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">I welcome an opportunity to further discuss my qualifications with you and learn more about the internship role(s) at your earliest convenience. Thank you very much for your time and consideration.</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">&nbsp;</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">Best Regards,</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><span style="font-family:&quot;Calibri Light&quot;,sans-serif">&nbsp;</span></p>
                    <p class="MsoNormal" style="margin:0cm 0cm 0.0001pt;line-height:normal;font-size:11pt;font-family:Calibri,sans-serif"><strong><span style="font-family:&quot;Calibri Light&quot;,sans-serif">Hrithik Shah</span></strong></p>
                    <span style="font-size:11pt;line-height:15.6933px;font-family:&quot;Calibri Light&quot;,sans-serif"><a href="http://hrithikshah.com/" rel="noopener" target="_blank" data-saferedirecturl="https://www.google.com/url?q=http://hrithikshah.com/&amp;source=gmail&amp;ust=1550005700791000&amp;usg=AFQjCNEeb9hMp3-juOdRunCjvEmufdarcA">hrithikshah.com</a></span>
                    <div class="yj6qo"></div>
                    <div class="adL">&nbsp;&nbsp;<br></div>
                 </div>
                 <div class="adL"></div>
              </div>
           </div>
           <div id=":15v" class="ii gt" style="display:none">
              <div id=":15u" class="a3s aXjCH undefined"></div>
           </div>
        </div>
          </body>
        </html>
        """

    part1 = MIMEText(html, "html")

    message.attach(part1)

    filename = "resume.pdf"  
    filenametwo = "coverletter.pdf"

    # Open PDF file in binary mode
    with open(filename, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        resume = MIMEBase("application", "octet-stream")
        resume.set_payload(attachment.read())

    with open(filenametwo, "rb") as attachment:
        # Add file as application/octet-stream
        # Email client can usually download this automatically as attachment
        coverletter = MIMEBase("application", "octet-stream")
        coverletter.set_payload(attachment.read())

    # Encode file in ASCII characters to send by email    
    encoders.encode_base64(resume)
    encoders.encode_base64(coverletter)

    # Add header as key/value pair to attachment part
    resume.add_header('Content-Disposition', 'attachment', filename=filename)

    coverletter.add_header('Content-Disposition', 'attachment', filename=filenametwo)

    # Add attachment to message and convert message to string
    message.attach(resume)
    message.attach(coverletter)
    text = message.as_string()

    # Create a secure SSL context
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", port, context=context) as server:
      server.login(sender_email, password)
      server.sendmail(sender_email, receiver_email, text)
      print ("Emailed "+firstName + " " + lastName + " at "+receiver_email)
    
  # updating excel sheet
  print("updating")
  #ws.cell(row=row+1, column=4).value = receiver_email
  ws.cell(row=row+1, column=11).value = "Emailed"
  ws.cell(row=row+1, column=12).value = datetime.now()
  wb.save(excelwb)

  book = xlrd.open_workbook(excelwb)
  sh = book.sheet_by_index(0)
  wb = openpyxl.load_workbook(filename = excelwb)
  ws = wb.worksheets[0]
  row += 1
  ws.cell(row=2, column=14).value = row

ws.cell(row=2, column=14).value = row
wb.save(excelwb)


print ("Done current emailing cycle")